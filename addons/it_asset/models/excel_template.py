# -*- coding: utf-8 -*-
"""
Excel Template Export untuk IT Asset Module
- Mengisi data Odoo ke template Excel yang sudah ada (tanpa mengubah layout)
- Support: Item Handover (BAST), Asset Request, Damage Report, dll
"""

import base64
import logging
import os
import tempfile

from odoo import api, fields, models, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    _logger.warning("openpyxl tidak terinstall. Install dengan: pip install openpyxl")
    openpyxl = None


class ITAssetExcelTemplate(models.AbstractModel):
    _name = 'it_asset.excel_template'
    _description = 'Excel Template Export untuk IT Asset'

    # ============================================
    # KONFIGURASI TEMPLATE
    # ============================================
    
    def _get_template_path(self, template_name):
        """
        Mendapatkan path file template Excel.
        Template diletakkan di: addons/it_asset/static/excel_templates/
        """
        module_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        template_dir = os.path.join(module_path, 'static', 'excel_templates')
        return os.path.join(template_dir, template_name)

    # ============================================
    # GENERIC: BUKA TEMPLATE & ISI DATA
    # ============================================

    def _load_template(self, template_name):
        """Membuka file template Excel"""
        if openpyxl is None:
            raise UserError(_(
                "Library 'openpyxl' tidak terinstall. "
                "Hubungi administrator untuk menginstallnya: pip install openpyxl"
            ))
        
        template_path = self._get_template_path(template_name)
        if not os.path.exists(template_path):
            raise UserError(_(
                "File template '%s' tidak ditemukan. "
                "Pastikan file template sudah ada di folder static/excel_templates/"
            ) % template_name)
        
        # Buka workbook langsung dari template (read-only copy via tempfile)
        wb = openpyxl.load_workbook(template_path)
        return wb

    def _save_workbook(self, wb):
        """
        Simpan workbook ke temporary file, baca sebagai base64.
        Tidak menggunakan BytesIO untuk menghindari potensi korupsi
        file yang mengandung gambar/objek bawaan template.
        """
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
            wb.save(tmp_path)
        
        with open(tmp_path, 'rb') as f:
            file_data = base64.b64encode(f.read())
        
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        
        wb.close()
        
        return file_data

    def _get_cell(self, ws, cell_ref):
        """Mendapatkan cell berdasarkan reference (contoh: 'B5')."""
        return ws[cell_ref]

    def _get_merged_cell_top_left(self, ws, cell_ref):
        """
        Jika cell_ref adalah bagian dari merged range, kembalikan referensi
        cell utama (top-left) dari merged range tersebut.
        """
        for merged_range in ws.merged_cells.ranges:
            if cell_ref in merged_range:
                top_left = merged_range.min_row, merged_range.min_col
                return f"{get_column_letter(top_left[1])}{top_left[0]}"
        return cell_ref

    def _set_cell_value(self, ws, cell_ref, value):
        """
        Mengisi nilai ke cell tertentu tanpa mengubah format.
        Handle merged cells untuk menghindari 'MergedCell' read-only error.
        """
        actual_ref = self._get_merged_cell_top_left(ws, cell_ref)
        cell = self._get_cell(ws, actual_ref)
        cell.value = value
        return cell

    # ============================================
    # 1. EXPORT ITEM HANDOVER (BAST) KE EXCEL
    # ============================================

    def export_item_handover_excel(self, handover_id):
        handover = self.env['it_asset.item.handover'].browse(handover_id)
        if not handover.exists():
            raise UserError(_("Item Handover tidak ditemukan!"))

        wb = self._load_template('bast_template.xlsx')
        ws = wb.active

        # Tanggal
        tgl = handover.handover_date
        tgl_str = tgl.strftime('%d/%m/%Y') if tgl else ''
        self._set_cell_value(ws, 'K11', tgl_str)
        
        # Pihak: Yang Menyerahkan
        sender_name = handover.sender_id.name if handover.sender_id else ''
        self._set_cell_value(ws, 'K7', sender_name)
        
        sender_job = handover.sender_id.job_id.name if handover.sender_id and handover.sender_id.job_id else ''
        self._set_cell_value(ws, 'K8', sender_job)
        
        # Pihak: Yang Menerima
        receiver_name = handover.receiver_id.name if handover.receiver_id else ''
        self._set_cell_value(ws, 'K9', receiver_name)
        
        receiver_job = handover.receiver_id.job_id.name if handover.receiver_id and handover.receiver_id.job_id else ''
        self._set_cell_value(ws, 'K10', receiver_job)

        # Tabel Items
        start_row = 14
        current_row = start_row
        
        for idx, line in enumerate(handover.line_ids, start=1):
            # No
            self._set_cell_value(ws, f'B{current_row}', idx)
            
            # Nama Barang
            if line.item_type == 'asset' and line.asset_id:
                item_name = line.asset_id.name if line.asset_id.name else ''
                if line.asset_id.asset_tag:
                    item_name += f" ({line.asset_id.asset_tag})"
            elif line.item_type == 'consumable' and line.consumable_id:
                item_name = line.consumable_id.name if line.consumable_id.name else ''
            else:
                item_name = ''
            self._set_cell_value(ws, f'D{current_row}', item_name)
            
            # Jumlah
            qty = line.quantity if line.quantity else 0
            self._set_cell_value(ws, f'S{current_row}', qty)
            
            # Kondisi
            if line.item_type == 'asset' and line.asset_id:
                kondisi = line.asset_id.condition if line.asset_id.condition else 'Baik'
            else:
                kondisi = 'Baik'
            self._set_cell_value(ws, f'X{current_row}', kondisi.capitalize())
            
            # Keterangan (SN / Notes)
            keterangan = ''
            if line.item_type == 'asset' and line.asset_id and line.asset_id.lot_id:
                keterangan += f"SN: {line.asset_id.lot_id.name}"
            if line.notes:
                if keterangan:
                    keterangan += '\n'
                keterangan += line.notes
            self._set_cell_value(ws, f'AF{current_row}', keterangan)
            
            current_row += 1

        file_data = self._save_workbook(wb)
        
        filename = f"BAST_{handover.name}_{tgl_str}.xlsx"
        filename = filename.replace('/', '_').replace('\\', '_')
        
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'datas': file_data,
            'res_model': 'it_asset.item.handover',
            'res_id': handover.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'name': _('Download Excel'),
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    # ============================================
    # 2. EXPORT ASSET REQUEST KE EXCEL
    # ============================================

    def export_asset_request_excel(self, request_id):
        request = self.env['it_asset.request'].browse(request_id)
        if not request.exists():
            raise UserError(_("Asset Request tidak ditemukan!"))

        wb = self._load_template('asset_request_template.xlsx')
        ws = wb.active

        self._set_cell_value(ws, 'C5', request.name if request.name else '')
        self._set_cell_value(ws, 'C6', request.employee_id.name if request.employee_id else '')
        self._set_cell_value(ws, 'C7', request.department_id.name if request.department_id else '')
        
        tgl = request.request_date
        self._set_cell_value(ws, 'C8', tgl.strftime('%d/%m/%Y') if tgl else '')
        
        self._set_cell_value(ws, 'C9', request.category_id.name if request.category_id else '')
        self._set_cell_value(ws, 'C10', request.reason if request.reason else '')
        
        state_label = dict(request._fields['state'].selection).get(request.state, '') if request.state else ''
        self._set_cell_value(ws, 'C11', state_label)

        file_data = self._save_workbook(wb)
        filename = f"Asset_Request_{request.name}.xlsx"

        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'datas': file_data,
            'res_model': 'it_asset.request',
            'res_id': request.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'name': _('Download Excel'),
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    # ============================================
    # 3. EXPORT DAMAGE REPORT KE EXCEL
    # ============================================

    def export_damage_report_excel(self, report_id):
        report = self.env['it_asset.damage_report'].browse(report_id)
        if not report.exists():
            raise UserError(_("Damage Report tidak ditemukan!"))

        wb = self._load_template('damage_report_template.xlsx')
        ws = wb.active

        self._set_cell_value(ws, 'C5', report.name if report.name else '')
        
        tgl = report.report_date
        self._set_cell_value(ws, 'C6', tgl.strftime('%d %B %Y') if tgl else '')
        
        self._set_cell_value(ws, 'C7', report.asset_id.name if report.asset_id else '')
        self._set_cell_value(ws, 'C8', report.asset_id.asset_tag if report.asset_id and report.asset_id.asset_tag else '')
        self._set_cell_value(ws, 'C9', report.employee_id.name if report.employee_id else '')
        self._set_cell_value(ws, 'C10', report.damage_type if report.damage_type else '')
        self._set_cell_value(ws, 'C11', report.description if report.description else '')

        file_data = self._save_workbook(wb)
        filename = f"Damage_Report_{report.name}.xlsx"

        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'datas': file_data,
            'res_model': 'it_asset.damage_report',
            'res_id': report.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'name': _('Download Excel'),
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    # ============================================
    # 4. EXPORT ACCOUNT REQUEST KE EXCEL
    # ============================================

    def export_account_request_excel(self, request_id):
        request = self.env['it_asset.account_request'].browse(request_id)
        if not request.exists():
            raise UserError(_("Account Request tidak ditemukan!"))

        wb = self._load_template('account_request_template.xlsx')
        ws = wb.active

        self._set_cell_value(ws, 'C5', request.name if request.name else '')
        self._set_cell_value(ws, 'C6', request.employee_id.name if request.employee_id else '')
        self._set_cell_value(ws, 'C7', request.department_id.name if request.department_id else '')
        
        tgl = request.request_date
        self._set_cell_value(ws, 'C8', tgl.strftime('%d/%m/%Y') if tgl else '')
        
        self._set_cell_value(ws, 'C9', request.account_type if request.account_type else '')
        self._set_cell_value(ws, 'C10', request.reason if request.reason else '')
        
        state_label = dict(request._fields['state'].selection).get(request.state, '') if request.state else ''
        self._set_cell_value(ws, 'C11', state_label)

        file_data = self._save_workbook(wb)
        filename = f"Account_Request_{request.name}.xlsx"

        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'datas': file_data,
            'res_model': 'it_asset.account_request',
            'res_id': request.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'name': _('Download Excel'),
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    # ============================================
    # 5. EXPORT HANDOVER (OLD) KE EXCEL
    # ============================================

    def export_handover_excel(self, handover_id):
        handover = self.env['it_asset.handover'].browse(handover_id)
        if not handover.exists():
            raise UserError(_("Handover tidak ditemukan!"))

        wb = self._load_template('handover_template.xlsx')
        ws = wb.active

        self._set_cell_value(ws, 'A9', handover.name if handover.name else '')
        
        tgl = handover.handover_date
        self._set_cell_value(ws, 'A11', tgl.strftime('%d %B %Y') if tgl else '')
        
        sender_name = handover.sender_id.name if handover.sender_id else ''
        self._set_cell_value(ws, 'I14', sender_name)
        
        sender_job = handover.sender_id.job_id.name if handover.sender_id and handover.sender_id.job_id else ''
        self._set_cell_value(ws, 'I15', sender_job)
        
        receiver_name = handover.receiver_id.name if handover.receiver_id else ''
        self._set_cell_value(ws, 'Z14', receiver_name)
        
        receiver_job = handover.receiver_id.job_id.name if handover.receiver_id and handover.receiver_id.job_id else ''
        self._set_cell_value(ws, 'Z15', receiver_job)
        
        self._set_cell_value(ws, 'D21', handover.asset_id.name if handover.asset_id else '')
        self._set_cell_value(ws, 'X21', handover.notes if handover.notes else '')

        file_data = self._save_workbook(wb)
        filename = f"Handover_{handover.name}.xlsx"

        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'datas': file_data,
            'res_model': 'it_asset.handover',
            'res_id': handover.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'name': _('Download Excel'),
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
